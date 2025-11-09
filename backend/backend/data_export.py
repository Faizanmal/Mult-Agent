"""
Data Export/Import Service for Enterprise Applications
Supports CSV, JSON, Excel, PDF exports
"""
import csv
import json
import io
from typing import List, Dict, Any
from datetime import datetime
from django.http import HttpResponse, StreamingHttpResponse
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logger = logging.getLogger(__name__)


class DataExporter:
    """
    Comprehensive data export service
    Supports multiple formats: CSV, JSON, Excel, PDF
    """
    
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = None) -> HttpResponse:
        """
        Export data to CSV format
        
        Args:
            data: List of dictionaries to export
            filename: Optional filename for download
        
        Returns:
            HttpResponse with CSV data
        """
        if not data:
            return HttpResponse("No data to export", status=400)
        
        filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Get headers from first item
        headers = list(data[0].keys())
        
        writer = csv.DictWriter(response, fieldnames=headers)
        writer.writeheader()
        
        for row in data:
            writer.writerow(row)
        
        logger.info(f"Exported {len(data)} rows to CSV: {filename}")
        return response
    
    @staticmethod
    def export_to_json(data: List[Dict[str, Any]], filename: str = None, 
                      pretty: bool = True) -> HttpResponse:
        """
        Export data to JSON format
        
        Args:
            data: List of dictionaries to export
            filename: Optional filename for download
            pretty: Whether to pretty-print JSON
        
        Returns:
            HttpResponse with JSON data
        """
        filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
        
        response = HttpResponse(content_type='application/json')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        export_data = {
            'exported_at': datetime.now().isoformat(),
            'count': len(data),
            'data': data
        }
        
        indent = 2 if pretty else None
        response.write(json.dumps(export_data, indent=indent, default=str))
        
        logger.info(f"Exported {len(data)} rows to JSON: {filename}")
        return response
    
    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str = None,
                       sheet_name: str = "Data") -> HttpResponse:
        """
        Export data to Excel format
        
        Args:
            data: List of dictionaries to export
            filename: Optional filename for download
            sheet_name: Name of the Excel sheet
        
        Returns:
            HttpResponse with Excel data
        """
        if not data:
            return HttpResponse("No data to export", status=400)
        
        filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Get headers
        headers = list(data[0].keys())
        
        # Style for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                ws.cell(row=row_num, column=col_num, value=str(value))
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        
        logger.info(f"Exported {len(data)} rows to Excel: {filename}")
        return response
    
    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], filename: str = None,
                     title: str = "Data Export") -> HttpResponse:
        """
        Export data to PDF format
        
        Args:
            data: List of dictionaries to export
            filename: Optional filename for download
            title: Title for the PDF document
        
        Returns:
            HttpResponse with PDF data
        """
        if not data:
            return HttpResponse("No data to export", status=400)
        
        filename = filename or f"export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Create PDF
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        
        # Add title
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
        elements.append(Paragraph("<br/><br/>", styles['Normal']))
        
        # Prepare table data
        headers = list(data[0].keys())
        table_data = [headers]
        
        for row in data:
            table_data.append([str(row.get(h, '')) for h in headers])
        
        # Create table
        table = Table(table_data)
        
        # Style table
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        logger.info(f"Exported {len(data)} rows to PDF: {filename}")
        return response
    
    @staticmethod
    def stream_large_export(queryset, format_type: str = 'csv', batch_size: int = 1000):
        """
        Stream large datasets efficiently
        
        Args:
            queryset: Django queryset to export
            format_type: Export format (csv, json)
            batch_size: Number of records per batch
        
        Yields:
            Chunks of exported data
        """
        def generate():
            if format_type == 'csv':
                # Get field names
                first_row = queryset.first()
                if not first_row:
                    return
                
                field_names = [f.name for f in first_row._meta.fields]
                
                # Write header
                output = io.StringIO()
                writer = csv.DictWriter(output, fieldnames=field_names)
                writer.writeheader()
                yield output.getvalue()
                
                # Write data in batches
                for i in range(0, queryset.count(), batch_size):
                    batch = queryset[i:i + batch_size]
                    output = io.StringIO()
                    writer = csv.DictWriter(output, fieldnames=field_names)
                    
                    for obj in batch:
                        row_data = {f: getattr(obj, f) for f in field_names}
                        writer.writerow(row_data)
                    
                    yield output.getvalue()
            
            elif format_type == 'json':
                yield '{"data": ['
                
                for i, obj in enumerate(queryset.iterator(chunk_size=batch_size)):
                    if i > 0:
                        yield ','
                    
                    obj_dict = {f.name: getattr(obj, f.name) for f in obj._meta.fields}
                    yield json.dumps(obj_dict, default=str)
                
                yield ']}'
        
        content_type = 'text/csv' if format_type == 'csv' else 'application/json'
        response = StreamingHttpResponse(generate(), content_type=content_type)
        response['Content-Disposition'] = f'attachment; filename="export.{format_type}"'
        
        return response


class DataImporter:
    """
    Data import service
    Supports CSV, JSON, Excel imports
    """
    
    @staticmethod
    def import_from_csv(file, model_class, field_mapping: Dict = None) -> Dict[str, Any]:
        """
        Import data from CSV file
        
        Args:
            file: Uploaded CSV file
            model_class: Django model to import into
            field_mapping: Optional mapping of CSV columns to model fields
        
        Returns:
            Import results with success/error counts
        """
        results = {
            'success': 0,
            'errors': 0,
            'error_details': []
        }
        
        try:
            # Read CSV
            decoded_file = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(decoded_file))
            
            for row_num, row in enumerate(csv_reader, 1):
                try:
                    # Map fields if mapping provided
                    if field_mapping:
                        row = {field_mapping.get(k, k): v for k, v in row.items()}
                    
                    # Create model instance
                    instance = model_class(**row)
                    instance.save()
                    
                    results['success'] += 1
                    
                except Exception as e:
                    results['errors'] += 1
                    results['error_details'].append({
                        'row': row_num,
                        'error': str(e),
                        'data': row
                    })
            
            logger.info(f"CSV import completed: {results['success']} success, {results['errors']} errors")
            
        except Exception as e:
            logger.error(f"CSV import failed: {str(e)}")
            results['error_details'].append({'error': str(e)})
        
        return results
    
    @staticmethod
    def import_from_json(file, model_class) -> Dict[str, Any]:
        """
        Import data from JSON file
        
        Args:
            file: Uploaded JSON file
            model_class: Django model to import into
        
        Returns:
            Import results
        """
        results = {
            'success': 0,
            'errors': 0,
            'error_details': []
        }
        
        try:
            data = json.load(file)
            
            # Handle both list and dict with 'data' key
            if isinstance(data, dict) and 'data' in data:
                data = data['data']
            
            for item_num, item in enumerate(data, 1):
                try:
                    instance = model_class(**item)
                    instance.save()
                    results['success'] += 1
                    
                except Exception as e:
                    results['errors'] += 1
                    results['error_details'].append({
                        'item': item_num,
                        'error': str(e),
                        'data': item
                    })
            
            logger.info(f"JSON import completed: {results['success']} success, {results['errors']} errors")
            
        except Exception as e:
            logger.error(f"JSON import failed: {str(e)}")
            results['error_details'].append({'error': str(e)})
        
        return results
